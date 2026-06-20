from __future__ import annotations

import json
import os
import shutil
import time
from contextlib import contextmanager
from pathlib import Path
from typing import Any, Iterator

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

from .config import (
    BACKUP_DIR,
    CONFIDENTIALITY,
    LOCK_PATH,
    PRIORITIES,
    SHEET_ORDER,
    STATUSES,
    WORKBOOK_PATH,
)
from .ids import next_id, utc_now
from .schema import SHEETS

NAVY = "183153"
TEAL = "2D6F73"
LIGHT_TEAL = "DDEEEF"
CREAM = "F6F2E9"
INK = "202A35"
MUTED = "667085"
WHITE = "FFFFFF"
GREEN = "D9EAD3"
AMBER = "FCE5CD"
RED = "F4CCCC"
LIGHT_BLUE = "EAF1F8"
THIN = Side(style="thin", color="D9E2E8")


@contextmanager
def workbook_lock(timeout: float = 10.0) -> Iterator[None]:
    start = time.monotonic()
    while True:
        try:
            fd = os.open(LOCK_PATH, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
            os.write(fd, f"{os.getpid()} {utc_now()}".encode())
            os.close(fd)
            break
        except FileExistsError:
            if time.monotonic() - start > timeout:
                raise TimeoutError("Workbook is locked by another Product Career Ops process")
            time.sleep(0.1)
    try:
        yield
    finally:
        Path(LOCK_PATH).unlink(missing_ok=True)


def _style_header(row) -> None:
    for cell in row:
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(color=WHITE, bold=True, size=10)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="medium", color=TEAL))


def _style_body(ws) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(color=INK, size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=THIN)


def _column_widths(ws) -> None:
    text_wide = {
        "Statement", "Evidence", "Actions", "Result", "Notes", "Rationale",
        "Unknowns", "Situation", "Initial_Recommendation", "Refined_Response",
        "Human_Judgment", "Final_Decision", "Career_Direction",
        "Next_Quarter_Bets", "Market_Thesis", "Risks", "Feedback",
    }
    medium = {
        "Company", "Role", "Recommendation", "Next_Action", "Context",
        "Problem", "Stakeholders", "Tradeoffs", "Reflection", "Competencies",
        "Topics", "Prompt_Lesson",
    }
    for idx, cell in enumerate(ws[1], 1):
        header = str(cell.value or "")
        if header in text_wide:
            width = 34
        elif header in medium:
            width = 24
        elif "Path" in header or "URL" in header:
            width = 28
        elif "Date" in header or header.endswith("_At"):
            width = 19
        elif header.endswith("_ID") or header == "Score_100":
            width = 14
        else:
            width = min(max(len(header) + 3, 12), 20)
        ws.column_dimensions[cell.column_letter].width = width


def _add_table(ws, name: str) -> None:
    end_col = ws.max_column
    end_row = max(ws.max_row, 2)
    if ws.max_row == 1:
        ws.append(["" for _ in range(end_col)])
    ref = f"A1:{ws.cell(end_row, end_col).coordinate}"
    table = Table(displayName=name.replace("_", "")[:25] + "Table", ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(table)


def _create_dashboard(wb) -> None:
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:H2")
    ws["A1"] = "PRODUCT CAREER OPS"
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].font = Font(color=WHITE, bold=True, size=22)
    ws["A1"].alignment = Alignment(vertical="center")

    ws.merge_cells("A4:D4")
    ws["A4"] = "JOB SEARCH"
    ws.merge_cells("E4:H4")
    ws["E4"] = "DEVELOPMENT"
    for cell in ("A4", "E4"):
        ws[cell].fill = PatternFill("solid", fgColor=TEAL)
        ws[cell].font = Font(color=WHITE, bold=True, size=13)
        ws[cell].alignment = Alignment(horizontal="center")

    cards = [
        ("A6", "Active opportunities", '=COUNTIF(Opportunities!I:I,"Active")'),
        ("C6", "Applications", '=COUNTA(Applications!A:A)-1'),
        ("A9", "Interviews", '=COUNTA(Interviews!A:A)-1'),
        ("C9", "Offers", '=COUNTIF(Opportunities!I:I,"Offer")'),
        ("E6", "Coaching sessions", '=COUNTA(Coaching_Sessions!A:A)-1'),
        ("G6", "Decisions", '=COUNTA(Decisions!A:A)-1'),
        ("E9", "Weekly reviews", '=COUNTA(Weekly_Reviews!A:A)-1'),
        ("G9", "Experiments", '=COUNTIF(Development_Experiments!H:H,"Active")'),
    ]
    for anchor, label, formula in cards:
        col = ws[anchor].column
        row = ws[anchor].row
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
        ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 2, end_column=col + 1)
        ws.cell(row, col).value = label
        ws.cell(row, col).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        ws.cell(row, col).font = Font(color=MUTED, bold=True, size=10)
        ws.cell(row + 1, col).value = formula
        ws.cell(row + 1, col).font = Font(color=NAVY, bold=True, size=24)
        ws.cell(row + 1, col).alignment = Alignment(horizontal="center", vertical="center")
        for r in range(row, row + 3):
            for c in range(col, col + 2):
                ws.cell(r, c).border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    ws.merge_cells("A14:H14")
    ws["A14"] = "Development builds evidence -> Search tests it -> Market feedback guides the next development bet"
    ws["A14"].fill = PatternFill("solid", fgColor=CREAM)
    ws["A14"].font = Font(color=INK, italic=True, size=11)
    ws["A14"].alignment = Alignment(horizontal="center")
    for col in "ABCDEFGH":
        ws.column_dimensions[col].width = 16
    for row in (1, 2):
        ws.row_dimensions[row].height = 24


def create_workbook(path: Path = WORKBOOK_PATH) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    _create_dashboard(wb)

    for sheet_name in SHEET_ORDER:
        if sheet_name == "Dashboard":
            continue
        ws = wb.create_sheet(sheet_name)
        headers = SHEETS.get(sheet_name, ["Key", "Value"])
        ws.append(headers)
        _style_header(ws[1])
        ws.row_dimensions[1].height = 34
        _column_widths(ws)
        if sheet_name != "Lists":
            _add_table(ws, sheet_name)

    lists = wb["Lists"]
    lists.delete_rows(2, lists.max_row)
    for list_name, values in (
        ("Status", STATUSES),
        ("Priority", PRIORITIES),
        ("Confidentiality", CONFIDENTIALITY),
    ):
        for idx, value in enumerate(values, 1):
            lists.append([list_name, value, idx])
    _style_body(lists)
    _add_table(lists, "Lists")

    _apply_workbook_formats(wb)
    wb.save(path)
    validate_workbook(path)
    return path


def _apply_workbook_formats(wb) -> None:
    for ws in wb.worksheets:
        if ws.title == "Dashboard":
            continue
        _style_body(ws)
        _column_widths(ws)

    profile = wb["Profile"]
    profile.column_dimensions["A"].width = 24
    profile.column_dimensions["B"].width = 80
    profile.column_dimensions["C"].width = 24

    opp = wb["Opportunities"]
    opp.conditional_formatting.add(
        "K2:K5000",
        CellIsRule(operator="greaterThanOrEqual", formula=["80"], fill=PatternFill("solid", fgColor=GREEN)),
    )
    opp.conditional_formatting.add(
        "K2:K5000",
        CellIsRule(operator="between", formula=["70", "79"], fill=PatternFill("solid", fgColor=AMBER)),
    )
    opp.conditional_formatting.add(
        "K2:K5000",
        CellIsRule(operator="lessThan", formula=["70"], fill=PatternFill("solid", fgColor=RED)),
    )
    opp.conditional_formatting.add(
        "I2:I5000",
        FormulaRule(formula=['$I2="Active"'], fill=PatternFill("solid", fgColor=GREEN)),
    )


def validate_workbook(path: Path = WORKBOOK_PATH) -> None:
    if not path.exists():
        raise FileNotFoundError(path)
    wb = load_workbook(path, read_only=True, data_only=False)
    missing = [name for name in SHEET_ORDER if name not in wb.sheetnames]
    if missing:
        raise ValueError(f"Workbook missing sheets: {', '.join(missing)}")
    for sheet, headers in SHEETS.items():
        ws = wb[sheet]
        actual = [ws.cell(1, col).value for col in range(1, len(headers) + 1)]
        if actual != headers:
            raise ValueError(f"{sheet} schema mismatch: {actual} != {headers}")
    wb.close()


def _backup() -> None:
    if not WORKBOOK_PATH.exists():
        return
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    stamp = utc_now().replace(":", "").replace("+00:00", "Z")
    shutil.copy2(WORKBOOK_PATH, BACKUP_DIR / f"product-career-ops-{stamp}.xlsx")


def _replace_atomically(wb) -> None:
    tmp = WORKBOOK_PATH.with_name(WORKBOOK_PATH.stem + ".tmp.xlsx")
    wb.save(tmp)
    validate_workbook(tmp)
    _backup()
    os.replace(tmp, WORKBOOK_PATH)


def _rows_as_dicts(ws) -> list[dict[str, Any]]:
    headers = [cell.value for cell in ws[1]]
    rows = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        if not any(value not in (None, "") for value in values):
            continue
        rows.append(dict(zip(headers, values)))
    return rows


def read_rows(sheet: str) -> list[dict[str, Any]]:
    validate_workbook()
    wb = load_workbook(WORKBOOK_PATH, read_only=True, data_only=False)
    rows = _rows_as_dicts(wb[sheet])
    wb.close()
    return rows


def append_record(
    sheet: str,
    data: dict[str, Any],
    *,
    actor: str = "pco",
    action: str = "create",
    entity_id: str | None = None,
    history: bool = True,
    summary: str | None = None,
) -> str:
    if sheet not in SHEETS:
        raise KeyError(sheet)
    with workbook_lock():
        if not WORKBOOK_PATH.exists():
            create_workbook()
        validate_workbook()
        wb = load_workbook(WORKBOOK_PATH)
        ws = wb[sheet]
        headers = SHEETS[sheet]
        id_header = headers[0] if headers[0].endswith("_ID") else None
        if id_header:
            current = [str(cell.value) for cell in ws["A"][1:] if cell.value]
            entity_id = entity_id or next_id(current, sheet)
            data[id_header] = entity_id
        values = [data.get(header, "") for header in headers]
        placeholder_row = (
            ws.max_row == 2
            and not any(ws.cell(2, col).value not in (None, "") for col in range(1, ws.max_column + 1))
        )
        if placeholder_row:
            for col, value in enumerate(values, 1):
                ws.cell(2, col).value = value
        else:
            ws.append(values)
        _style_body(ws)
        if ws.tables:
            table = next(iter(ws.tables.values()))
            table.ref = f"A1:{ws.cell(ws.max_row, ws.max_column).coordinate}"
        if history and sheet != "History":
            _append_history_in_workbook(
                wb, actor, action, sheet, entity_id or "",
                summary or data.get("Summary", action), "", data
            )
        _apply_workbook_formats(wb)
        _replace_atomically(wb)
        wb.close()
        return entity_id or ""


def _append_history_in_workbook(
    wb, actor: str, action: str, entity_type: str, entity_id: str,
    summary: str, before: Any, after: Any,
) -> None:
    ws = wb["History"]
    current = [str(cell.value) for cell in ws["A"][1:] if cell.value]
    event_id = next_id(current, "History")
    ws.append([
        event_id,
        utc_now(),
        actor,
        action,
        entity_type,
        entity_id,
        summary,
        json.dumps(before, default=str, ensure_ascii=False) if before != "" else "",
        json.dumps(after, default=str, ensure_ascii=False) if after != "" else "",
    ])
    if ws.tables:
        table = next(iter(ws.tables.values()))
        table.ref = f"A1:{ws.cell(ws.max_row, ws.max_column).coordinate}"


def find_record(sheet: str, entity_id: str) -> dict[str, Any] | None:
    for row in read_rows(sheet):
        first = next(iter(row.values()))
        if first == entity_id:
            return row
    return None


def counts() -> dict[str, int]:
    return {sheet: len(read_rows(sheet)) for sheet in SHEETS if sheet != "Lists"}
