from pathlib import Path

from openpyxl import load_workbook

from pco.config import SHEET_ORDER
from pco.workbook import create_workbook, validate_workbook


def test_workbook_schema(tmp_path: Path):
    path = tmp_path / "test.xlsx"
    create_workbook(path)
    validate_workbook(path)
    wb = load_workbook(path, read_only=True)
    assert wb.sheetnames == SHEET_ORDER
    assert wb["Opportunities"]["A1"].value == "Opportunity_ID"
    assert wb["Coaching_Sessions"]["A1"].value == "Coaching_ID"
    wb.close()

